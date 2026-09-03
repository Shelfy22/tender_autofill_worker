from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "N",
    "Название товара тендера",
    "Количество",
    "Ед. изм.",
    "Артикул",
    "Ссылка",
    "Код ETM",
    "Наименование ETM",
    "Производитель",
    "Медианная цена",
    "Валюта",
    "Сумма позиции",
    "Соответствие",
    "Обоснование",
    "Источник",
]


def _value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _source_label(detail: dict[str, Any]) -> str:
    reference = detail.get("sourceReference")
    if not isinstance(reference, dict):
        return ""
    parts = [
        str(reference.get("fileName") or "").strip(),
        str(reference.get("sheet") or "").strip(),
        f"row {reference.get('row')}" if reference.get("row") else "",
    ]
    return " / ".join(part for part in parts if part)


def build_product_matching_workbook(product_check: dict[str, Any]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Автоподбор"
    worksheet.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    details = product_check.get("details")
    if not isinstance(details, list):
        details = []

    for index, detail in enumerate(details, start=1):
        if not isinstance(detail, dict):
            continue
        result = detail.get("result") if isinstance(detail.get("result"), dict) else {}
        article = detail.get("article") or result.get("Артикул") or ""
        link = detail.get("link") or result.get("Ссылка") or ""
        product_id = detail.get("productId") or result.get("ID товара") or article
        worksheet.append(
            [
                index,
                _value(detail.get("sourceProduct") or detail.get("productQuery")),
                _value(detail.get("quantity")),
                _value(detail.get("unit")),
                _value(article),
                _value(link),
                _value(product_id),
                _value(result.get("Наименование") or ""),
                _value(result.get("Производитель") or ""),
                _value(detail.get("medianUnitPriceRub")),
                _value(detail.get("priceCurrency") or result.get("Валюта")),
                _value(detail.get("positionTotalPriceRub")),
                _value(result.get("Соответствие") or ""),
                _value(result.get("Обоснование") or ""),
                _source_label(detail),
            ]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = [8, 48, 14, 12, 16, 34, 16, 52, 24, 16, 10, 16, 22, 72, 40]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()