from __future__ import annotations

import codecs
import csv
import io
import subprocess
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.config import Settings
from app.models import SpreadsheetRow, SpreadsheetTable
from app.services.parsers.archives import validate_zip_container


KEYWORDS = (
    "наименование", "товар", "работ", "услуг", "характеристик", "кол-во", "количество",
    "ед. изм", "цена", "стоимость", "нмц", "спецификац", "поставка", "предмет закупки",
    "окпд", "производител", "страна происхождения",
)


def _row_line(number: int, values: Iterable[object]) -> str:
    cells = [str(value).replace("\n", " ").strip() for value in values if value is not None and str(value).strip()]
    return f"Строка {number}: " + " | ".join(cells) if cells else ""


def _worksheet_row_line(number: int, values: Iterable[object]) -> str:
    # Keep the Excel column address. Dropping empty cells without addresses shifts
    # the apparent schema and was one source of incorrect Extract-to-JSON results.
    cells = []
    for column, value in enumerate(values, start=1):
        if value is None or not str(value).strip():
            continue
        normalized = str(value).replace("\r", " ").replace("\n", " ").strip()
        cells.append(f"{get_column_letter(column)}: {normalized}")
    return f"Строка {number}: " + " | ".join(cells) if cells else ""


def _xlsx_text(path: Path, settings: Settings) -> str:
    validate_zip_container(path, settings)
    # read_only streams worksheet rows; data_only returns cached formula results
    # instead of formula source text. keep_links=False avoids retaining external-link data.
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    lines: list[str] = []
    total = 0
    try:
        for sheet in workbook.worksheets:
            header = f"Лист: {sheet.title}"
            lines.append(header)
            total += len(header)
            for number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                line = _worksheet_row_line(number, row)
                if not line:
                    continue
                remaining = settings.max_text_chars_per_file - total
                if remaining <= 0:
                    return "\n".join(lines)
                lines.append(line[:remaining])
                total += min(len(line), remaining) + 1
    finally:
        workbook.close()
    return "\n".join(lines)


def _xlsx_tables(
    path: Path,
    settings: Settings,
) -> tuple[str, list[SpreadsheetTable], list[str]]:
    validate_zip_container(path, settings)
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    tables: list[SpreadsheetTable] = []
    warnings: list[str] = []
    lines: list[str] = []
    structured_chars = 0
    limit_reached = False
    try:
        for sheet in workbook.worksheets:
            sheet_header = f"Лист: {sheet.title}"
            lines.append(sheet_header)
            structured_chars += len(sheet_header)
            rows: list[SpreadsheetRow] = []
            sheet_warnings: list[str] = []
            for row_number, values in enumerate(
                sheet.iter_rows(values_only=True),
                start=1,
            ):
                cells: dict[str, str] = {}
                try:
                    for column_number, raw_value in enumerate(values, start=1):
                        if raw_value is None:
                            continue
                        value = (
                            str(raw_value)
                            .replace(chr(13), " ")
                            .replace(chr(10), " ")
                            .strip()
                        )
                        if not value:
                            continue
                        remaining = settings.max_text_chars_per_file - structured_chars
                        if remaining <= 0:
                            limit_reached = True
                            break
                        value = value[: min(4000, remaining)]
                        cells[get_column_letter(column_number)] = value
                        structured_chars += len(value)
                    if cells:
                        rows.append(SpreadsheetRow(row=row_number, cells=cells))
                        lines.append(
                            f"Строка {row_number}: "
                            + " | ".join(
                                f"{column}: {value}"
                                for column, value in cells.items()
                            )
                        )
                except Exception as exc:
                    message = (
                        f"Лист {sheet.title}, строка {row_number}: "
                        f"структурная строка пропущена: {type(exc).__name__}: {exc}"
                    )
                    sheet_warnings.append(message)
                    warnings.append(message)
                if limit_reached:
                    break

            try:
                from app.services.products import infer_spreadsheet_headers

                header_rows, header_map, header_labels = infer_spreadsheet_headers(rows)
                tables.append(
                    SpreadsheetTable(
                        fileName=path.name,
                        sheet=sheet.title,
                        headerRows=header_rows,
                        headerMap=header_map,
                        headerLabels=header_labels,
                        rows=rows,
                        parserWarnings=sheet_warnings,
                    )
                )
            except Exception as exc:
                warnings.append(
                    f"Лист {sheet.title}: структурная таблица не создана: "
                    f"{type(exc).__name__}: {exc}. Используется текстовый fallback."
                )
            if limit_reached:
                warnings.append(
                    "Структурное представление Excel обрезано по лимиту символов; "
                    "оставшаяся часть доступна через текстовый fallback в пределах общего лимита."
                )
                break
    finally:
        workbook.close()
    return chr(10).join(lines), tables, warnings


def _csv_text(path: Path, settings: Settings) -> str:
    decoder = codecs.getincrementaldecoder("utf-8")("strict")
    utf8_valid = True
    try:
        with path.open("rb") as binary:
            while chunk := binary.read(1024 * 1024):
                decoder.decode(chunk)
            decoder.decode(b"", final=True)
    except UnicodeDecodeError:
        utf8_valid = False
    encoding = "utf-8-sig" if utf8_valid else "cp1251"
    with path.open("rb") as binary:
        sample = binary.read(10_000).decode(encoding, errors="replace")
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";\t,")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"
    lines: list[str] = []
    total = 0
    with path.open("rb") as binary, io.TextIOWrapper(
        binary, encoding=encoding, errors="replace", newline=""
    ) as text_stream:
        for number, row in enumerate(csv.reader(text_stream, delimiter=delimiter), start=1):
            line = _worksheet_row_line(number, row)
            remaining = settings.max_text_chars_per_file - total
            if remaining <= 0:
                break
            if line:
                lines.append(line[:remaining])
                total += min(len(line), remaining) + 1
    return "\n".join(lines)


def _convert_xls(path: Path, settings: Settings) -> Path:
    output_dir = path.parent / "converted"
    output_dir.mkdir(exist_ok=True)
    subprocess.run(
        [settings.libreoffice_binary, "--headless", "--convert-to", "xlsx", "--outdir", str(output_dir), str(path)],
        capture_output=True,
        timeout=settings.conversion_timeout_seconds,
        check=True,
    )
    result = output_dir / f"{path.stem}.xlsx"
    if not result.exists():
        raise RuntimeError("LibreOffice не создал XLSX из XLS")
    return result


def extract_spreadsheet_content(
    path: Path,
    file_type: str,
    settings: Settings,
) -> tuple[str, str, list[str], list[SpreadsheetTable]]:
    warnings: list[str] = []
    tables: list[SpreadsheetTable] = []
    if file_type == "csv":
        text = _csv_text(path, settings)
    else:
        source = _convert_xls(path, settings) if file_type == "xls" else path
        if file_type == "xls":
            warnings.append("Старый XLS преобразован LibreOffice в XLSX.")
        text, structured_tables, structured_warnings = _xlsx_tables(
            source,
            settings,
        )
        tables.extend(structured_tables)
        warnings.extend(structured_warnings)
    lowered = text.lower()
    useful = len(text) >= 80 and any(keyword in lowered for keyword in KEYWORDS)
    if not useful:
        warnings.append(
            f"Таблица извлечена, но не прошла keyword quality check: длина {len(text)}. "
            "Текст сохранён для LLM и детерминированного разбора."
        )
    return text, ("ok" if useful else "spreadsheet_low_confidence"), warnings, tables


def extract_spreadsheet_text(
    path: Path,
    file_type: str,
    settings: Settings,
) -> tuple[str, str, list[str]]:
    text, status, warnings, _ = extract_spreadsheet_content(
        path,
        file_type,
        settings,
    )
    return text, status, warnings
