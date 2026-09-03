import json
from pathlib import Path

from openpyxl import Workbook

from app.config import Settings
from app.models import ParsedDocument, SpreadsheetRow, SpreadsheetTable
from app.services.documents import DocumentProcessor
from app.services.parsers.spreadsheets import (
    extract_spreadsheet_content,
    extract_spreadsheet_text,
)
from app.services.products import extract_deterministic_positions


def test_xlsx_streaming_text_preserves_column_addresses_and_quantity(tmp_path: Path) -> None:
    path = tmp_path / "specification.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Спецификация"
    sheet.append(["№ п/п", "Наименование товара", "Характеристика", "Ед. изм.", "Количество"])
    sheet.append([1, "Кабель силовой", None, "шт", 12])
    workbook.save(path)

    settings = Settings(
        postgres_dsn="postgresql://user:pass@localhost/db",
        max_text_chars_per_file=10_000,
    )
    text, status, warnings = extract_spreadsheet_text(path, "xlsx", settings)

    assert status == "ok"
    assert not warnings
    assert "Лист: Спецификация" in text
    assert "Строка 2: A: 1 | B: Кабель силовой | D: шт | E: 12" in text
    positions = extract_deterministic_positions(text)
    assert [(item.product, item.unit, item.quantity) for item in positions] == [
        ("Кабель силовой", "шт", 12.0)
    ]


def test_xlsx_structured_json_has_header_map_and_valid_rows(tmp_path: Path) -> None:
    path = tmp_path / "structured.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары"
    sheet.append(
        [
            "№",
            "Наименование товара",
            "Ед. изм.",
            "Количество",
            "Количество полюсов",
        ]
    )
    sheet.append([1, "Выключатель автоматический", "шт", 10, 3])
    workbook.save(path)

    settings = Settings(
        postgres_dsn="postgresql://user:pass@localhost/db",
        max_text_chars_per_file=10_000,
    )
    text, status, warnings, tables = extract_spreadsheet_content(
        path,
        "xlsx",
        settings,
    )

    assert status == "ok"
    assert not warnings
    assert len(tables) == 1
    table = tables[0]
    assert table.headerMap["product"] == "B"
    assert table.headerMap["unit"] == "C"
    assert table.headerMap["quantity"] == "D"
    assert table.headerRows == [1]
    assert table.rows[1].cells == {
        "A": "1",
        "B": "Выключатель автоматический",
        "C": "шт",
        "D": "10",
        "E": "3",
    }

    serialized = json.dumps(
        [item.model_dump() for item in tables],
        ensure_ascii=False,
    )
    restored = [
        SpreadsheetTable.model_validate(item)
        for item in json.loads(serialized)
    ]
    positions = extract_deterministic_positions(text, restored)

    assert len(positions) == 1
    assert positions[0].product == "Выключатель автоматический"
    assert positions[0].quantity == 10
    assert positions[0].sourceCells["D"] == "10"
    assert positions[0].sourceCells["E"] == "3"


def test_document_processor_keeps_structured_tables_in_parsed_document(
    tmp_path: Path,
) -> None:
    path = tmp_path / "document.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Спецификация"
    sheet.append(["Наименование товара", "Ед. изм.", "Количество"])
    sheet.append(["Кабель силовой", "м", 25])
    workbook.save(path)
    settings = Settings(
        postgres_dsn="postgresql://user:pass@localhost/db",
        temp_root=tmp_path,
        max_text_chars_per_file=10_000,
    )
    processor = DocumentProcessor(settings, tmp_path)
    try:
        documents = processor._process_path(
            path,
            {"index": 1, "fileName": "document.xlsx"},
            depth=0,
        )
    finally:
        processor.close()

    assert len(documents) == 1
    assert documents[0].parserStatus == "ok"
    assert len(documents[0].spreadsheetTables) == 1
    assert documents[0].spreadsheetTables[0].fileName == "document.xlsx"
    positions = extract_deterministic_positions(
        documents[0].text,
        documents[0].spreadsheetTables,
    )
    assert [(item.product, item.quantity, item.unit) for item in positions] == [
        ("Кабель силовой", 25.0, "м")
    ]


def test_low_confidence_xlsx_text_is_not_discarded(tmp_path: Path) -> None:
    path = tmp_path / "short.xlsx"
    workbook = Workbook()
    workbook.active.append(["ABC", 123])
    workbook.save(path)

    settings = Settings(
        postgres_dsn="postgresql://user:pass@localhost/db",
        max_text_chars_per_file=10_000,
    )
    text, status, warnings = extract_spreadsheet_text(path, "xlsx", settings)

    assert "A: ABC | B: 123" in text
    assert status == "spreadsheet_low_confidence"
    assert warnings


def test_xlsx_prices_are_extracted_by_headers_with_source_coordinates(tmp_path: Path) -> None:
    path = tmp_path / "priced_specification.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Спецификация"
    sheet.append(
        [
            "№ п/п",
            "Наименование товара",
            "Ед. изм.",
            "Количество",
            "Цена за единицу, руб.",
            "Стоимость позиции, руб.",
        ]
    )
    sheet.append([1, "Стеллаж Универсал", "шт", 10, "15 707,43", "157 074,30"])
    workbook.save(path)

    settings = Settings(
        postgres_dsn="postgresql://user:pass@localhost/db",
        max_text_chars_per_file=10_000,
    )
    text, status, warnings = extract_spreadsheet_text(path, "xlsx", settings)
    positions = extract_deterministic_positions(
        "--- ДОКУМЕНТ 1 ---\nfileName: priced_specification.xlsx\n" + text
    )

    assert status == "ok"
    assert not warnings
    assert len(positions) == 1
    position = positions[0]
    assert position.documentUnitPriceRub == 15_707.43
    assert position.documentLineTotalRub == 157_074.30
    assert position.documentCurrency == "RUB"
    assert position.documentPriceSource is not None
    assert position.documentPriceSource.fileName == "priced_specification.xlsx"
    assert position.documentPriceSource.sheet == "Спецификация"
    assert position.documentPriceSource.row == 2
    assert position.documentPriceSource.unitPriceColumn == "E"
    assert position.documentPriceSource.lineTotalColumn == "F"
    assert position.documentPriceSource.extractionMethod == "excel_deterministic"
    assert position.sourceReference is not None
    assert position.sourceReference.fileName == "priced_specification.xlsx"
    assert position.sourceReference.sheet == "Спецификация"
    assert position.sourceReference.row == 2
    assert position.sourceReference.productColumn == "B"
    assert position.sourceReference.unitColumn == "C"
    assert position.sourceReference.quantityColumn == "D"
    assert position.sourceReference.extractionMethod == "excel_deterministic"

def test_product_extraction_text_uses_only_spreadsheet_documents() -> None:
    from app.pipeline import build_product_extraction_text

    spreadsheet = ParsedDocument(
        documentIndex=1,
        fileName="spec.xlsx",
        fileExtension="xlsx",
        parserStatus="ok",
        text="Лист: Sheet1\nСтрока 1: A: Наименование товара | B: Количество\nСтрока 2: A: Cable | B: 10",
        spreadsheetTables=[
            SpreadsheetTable(
                sheet="Sheet1",
                rows=[SpreadsheetRow(row=2, cells={"A": "Cable", "B": "10"})],
            )
        ],
    )
    contract = ParsedDocument(
        documentIndex=2,
        fileName="contract.docx",
        fileExtension="docx",
        parserStatus="ok",
        text="Contract text should not be sent to product extraction",
    )

    text, warnings, spreadsheet_only = build_product_extraction_text(
        "combined text with contract",
        [spreadsheet, contract],
    )

    assert spreadsheet_only is True
    assert warnings
    assert "Cable" in text
    assert "contract.docx" not in text
    assert "Contract text should not be sent" not in text


def test_product_extraction_text_falls_back_to_combined_text_without_spreadsheets() -> None:
    from app.pipeline import build_product_extraction_text

    text, warnings, spreadsheet_only = build_product_extraction_text(
        "combined text with pdf products",
        [
            ParsedDocument(
                documentIndex=1,
                fileName="spec.pdf",
                fileExtension="pdf",
                parserStatus="ok",
                text="pdf text",
            )
        ],
    )

    assert text == "combined text with pdf products"
    assert warnings == []
    assert spreadsheet_only is False
